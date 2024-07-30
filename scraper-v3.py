import argparse
import requests
from bs4 import BeautifulSoup
import pandas as pd
from lxml import etree
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
from datetime import datetime
import os
import time
import re

def get_all_pages_from_sitemaps(sitemap_urls):
    urls = []
    for sitemap_url in sitemap_urls:
        response = requests.get(sitemap_url)
        sitemap_xml = response.content
        root = etree.fromstring(sitemap_xml)
        namespace = {'ns': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
        urls.extend([url.text for url in root.xpath('//ns:loc', namespaces=namespace)])
    return urls

def sanitize_text(text):
    if isinstance(text, list):
        return [sanitize_text(t) for t in text]
    if text:
        return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text.replace('\n', ' ').replace('\r', '').strip())
    return text

def extract_html_structure(soup):
    structure = {
        'title': sanitize_text(soup.title.string if soup.title else 'No title'),
        'h1': sanitize_text([h1.get_text() for h1 in soup.find_all('h1')]),
        'h2': sanitize_text([h2.get_text() for h2 in soup.find_all('h2')]),
        'h3': sanitize_text([h3.get_text() for h3 in soup.find_all('h3')]),
        'h4': sanitize_text([h4.get_text() for h4 in soup.find_all('h4')]),
        'h5': sanitize_text([h5.get_text() for h5 in soup.find_all('h5')]),
        'h6': sanitize_text([h6.get_text() for h6 in soup.find_all('h6')]),
        'paragraphs': sanitize_text([p.get_text() for p in soup.find_all('p')]),
        'images': sanitize_text([img['src'] for img in soup.find_all('img', src=True)]),
        'links': sanitize_text([a['href'] for a in soup.find_all('a', href=True)]),
        'link_texts': sanitize_text([a.get_text() or 'Image: ' + (a.find('img')['alt'] if a.find('img') else 'No text') for a in soup.find_all('a', href=True)])
    }

    description = ''
    keywords = ''
    meta_tags = soup.find_all('meta')
    for tag in meta_tags:
        if tag.get('name') == 'description':
            description = tag.get('content')
        elif tag.get('name') == 'keywords':
            keywords = tag.get('content')

    structure['description'] = sanitize_text(description)
    structure['keywords'] = sanitize_text(keywords)

    return structure

def scrape_page(url, driver):
    driver.get(url)
    WebDriverWait(driver, 500).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    # WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".privy")))

    dataLayer = driver.execute_script("return window.dataLayer;")

    soup = BeautifulSoup(driver.page_source, 'html.parser')

    # Remove links from global elements
    if soup.header:
        for a in soup.header.find_all('a'):
            a.decompose()
    if soup.footer:
        for a in soup.footer.find_all('a'):
            a.decompose()
    for div in soup.find_all('div', class_='copyright'):
        for a in div.find_all('a'):
            a.decompose()
    for div in soup.find_all('div', class_='header'):
        for a in div.find_all('a'):
            a.decompose()
    oc_lcw_container = soup.find('div', id='oc-lcw-container')
    if oc_lcw_container:
        for a in oc_lcw_container.find_all('a'):
            a.decompose()
    for div in soup.find_all('div', class_='osano-cm-widget'):
        for a in div.find_all('a'):
            a.decompose()

    structure = extract_html_structure(soup)
    structure['url'] = sanitize_text(url)

    dataLayer_info = {
        'currency': None,
        'affiliation': None,
        'item_category': None,
        'item_category2': None,
        'item_id': None,
        'item_list_name': None,
        'item_name': None,
        'item_stock_status': None,
        'price': None
    }
    if dataLayer:
        for entry in dataLayer:
            if 'pageType' in entry and entry['pageType'] == 'product':
                ecommerce = entry.get('ecommerce', {})
                items = ecommerce.get('items', [])
                if items:
                    item = items[0]
                    dataLayer_info = {
                        'currency': ecommerce.get('currency'),
                        'affiliation': item.get('affiliation'),
                        'item_category': item.get('item_category'),
                        'item_category2': item.get('item_category2'),
                        'item_id': item.get('item_id'),
                        'item_list_name': item.get('item_list_name'),
                        'item_name': item.get('item_name'),
                        'item_stock_status': item.get('item_stock_status'),
                        'price': item.get('price')
                    }
                break

    structure.update(dataLayer_info)
    return structure

def capture_screenshot(driver, filename):
    # Set the window size to emulate desktop and capture the full page
    driver.set_window_size(1920, 1080)
    total_height = driver.execute_script("return document.body.scrollHeight")
    driver.set_window_size(1920, total_height)
    driver.find_element(By.TAG_NAME, 'body').screenshot(filename)

def sanitize_sheet_name(name):
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        name = name.replace(char, '')
    return name[:31]  # Excel sheet names must be <= 31 characters

def get_slug_from_url(url):
    parsed_url = urlparse(url)
    slug = parsed_url.path.strip('/').replace('/', '_')
    slug = slug.replace('en-us_', '').replace('/', '')
    return slug if slug else 'root'

def main(sitemap_urls, output_file, debug_mode=False, debug_limit=5, debug_urls=None):
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    screenshot_dir = f"screenshots-{timestamp}"
    os.makedirs(screenshot_dir, exist_ok=True)

    all_pages = get_all_pages_from_sitemaps(sitemap_urls)
    
    if debug_mode:
        if debug_urls:
            all_pages = debug_urls
        else:
            all_pages = all_pages[:debug_limit]

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    service = Service('/opt/homebrew/bin/chromedriver')
    driver = webdriver.Chrome(service=service, options=chrome_options)

    toc_entries = []
    
    workbook = Workbook()
    toc_ws = workbook.active
    toc_ws.title = "Table of Contents"

    # Add column headings to the Table of Contents
    toc_ws.cell(row=1, column=1, value="Sheet Name")
    toc_ws.cell(row=1, column=2, value="Page URL")

    for url in all_pages:
        try:
            page_data = scrape_page(url, driver)
            screenshot_filename = os.path.join(screenshot_dir, f"screenshot_{time.time()}.png")
            capture_screenshot(driver, screenshot_filename)
            
            df = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in page_data.items() if k not in ['url', 'links', 'link_texts']]))
            sheet_name = sanitize_sheet_name(get_slug_from_url(url))
            
            # Handle potential duplicates in sheet names
            original_sheet_name = sheet_name
            i = 1
            while sheet_name in workbook.sheetnames:
                sheet_name = f"{original_sheet_name}_{i}"
                i += 1
            
            # Add DataFrame to new sheet
            df_sheet = workbook.create_sheet(title=sheet_name)
            for r in dataframe_to_rows(df, index=False, header=True):
                df_sheet.append(r)

            # Add the Page URL and links to the last columns
            df_sheet['A1'] = 'Page URL'
            df_sheet['A2'] = page_data['url']
            links_col = len(df.columns) + 2  # Adding 2 to account for the Page URL column and header
            link_text_col = links_col + 1  # Next column for link texts

            df_sheet.cell(row=1, column=links_col, value='Links')
            df_sheet.cell(row=1, column=link_text_col, value='Link Text')
            for i, (link, link_text) in enumerate(zip(page_data['links'], page_data['link_texts']), start=2):
                df_sheet.cell(row=i, column=links_col, value=link)
                df_sheet.cell(row=i, column=link_text_col, value=link_text)
            
            # Insert the screenshot
            img = Image(screenshot_filename)
            df_sheet.add_image(img, f"A{len(df) + 20}")  # Place the image below the data

            # Add entry to TOC
            toc_entries.append((sheet_name, page_data['url']))

        except Exception as e:
            print(f"Error scraping {url}: {e}")
            toc_entries.append((url, f"Error: {e}"))

    # Create Table of Contents
    toc_df = pd.DataFrame(toc_entries, columns=["Sheet Name", "Page URL"])
    for index, row in toc_df.iterrows():
        sheet_name = row['Sheet Name']
        toc_ws.cell(row=index + 2, column=1).value = sheet_name
        toc_ws.cell(row=index + 2, column=1).hyperlink = f"=HYPERLINK#'{sheet_name}'!A1"
        toc_ws.cell(row=index + 2, column=2).value = str(row['Page URL'])

    workbook.save(output_file)
    driver.quit()
    print(f'Data saved to {output_file}')

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Scrape website data and save to Excel")
    parser.add_argument('sitemap_urls', type=str, nargs='+', help='URLs of the sitemaps')
    parser.add_argument('output_file', type=str, help='Name of the output Excel file')
    parser.add_argument('--debug', action='store_true', help='Activate debug mode')
    parser.add_argument('--limit', type=int, default=5, help='Limit the number of pages to scrape in debug mode')
    parser.add_argument('--urls', nargs='*', help='Specific URLs to scrape in debug mode')
    args = parser.parse_args()

    main(args.sitemap_urls, args.output_file, debug_mode=args.debug, debug_limit=args.limit, debug_urls=args.urls)